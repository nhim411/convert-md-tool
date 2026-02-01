"""
Excel Cleaner Module
Provides utility functions to clean and optimize Excel files before conversion.
Mainly handles 'Forward Fill' for merged cells to preserve context for RAG.
"""

import logging
import os
import shutil
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.utils import range_boundaries

logger = logging.getLogger(__name__)

def clean_excel_file(file_path: str) -> Optional[str]:
    """
    Clean Excel file by unmerging cells and filling values (Forward Fill).

    Args:
        file_path: Path to source .xlsx file

    Returns:
        Path to temporary cleaned file, or None if failed.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        modified = False

        for sheet in wb.worksheets:
            # List of merged ranges (copy to avoid modification during iteration)
            merged_ranges = list(sheet.merged_cells.ranges)

            if not merged_ranges:
                continue

            modified = True
            for merged_range in merged_ranges:
                # Get boundaries
                min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))

                # Get value of top-left cell
                top_left_value = sheet.cell(row=min_row, column=min_col).value

                # Unmerge
                sheet.unmerge_cells(str(merged_range))

                # Fill all cells in range with the value
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        cell.value = top_left_value

        if modified:
            # Save to temp file
            dir_name = os.path.dirname(file_path)
            base_name = os.path.basename(file_path)
            temp_name = f"cleaned_{base_name}"
            temp_path = os.path.join(dir_name, temp_name)

            wb.save(temp_path)
            wb.close()
            logger.info(f"Cleaned Excel file created: {temp_path}")
            return temp_path
        else:
            wb.close()
            return None # No changes needed

    except Exception as e:
        logger.error(f"Failed to clean Excel file {file_path}: {e}")
        return None
