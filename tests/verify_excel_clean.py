
import os
import sys
import openpyxl
from openpyxl.styles import Alignment

# Add app to path
sys.path.insert(0, os.path.join(os.getcwd(), 'app'))

import excel_cleaner

def create_test_file(filename="test_merged.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Create header
    ws['A1'] = "Category"
    ws['B1'] = "Item"
    ws['C1'] = "Value"

    # Create merged data
    # Category A (Rows 2-4)
    ws['A2'] = "Category A"
    ws.merge_cells('A2:A4')

    ws['B2'] = "Item 1"
    ws['B3'] = "Item 2"
    ws['B4'] = "Item 3"

    ws['C2'] = 10
    ws['C3'] = 20
    ws['C4'] = 30

    # Category B (Rows 5-6)
    ws['A5'] = "Category B"
    ws.merge_cells('A5:A6')

    ws['B5'] = "Item 4"
    ws['B6'] = "Item 5"

    wb.save(filename)
    print(f"Created test file: {filename}")
    return filename

def verify_cleaned(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Check A3 and A4 (should be "Category A")
    val_a3 = ws['A3'].value
    val_a4 = ws['A4'].value

    print(f"A3: {val_a3}")
    print(f"A4: {val_a4}")

    if val_a3 == "Category A" and val_a4 == "Category A":
        print("PASS: Category A filled down correctly.")
    else:
        print("FAIL: Category A not filled.")

    # Check A6
    val_a6 = ws['A6'].value
    print(f"A6: {val_a6}")
    if val_a6 == "Category B":
        print("PASS: Category B filled down correctly.")
    else:
        print("FAIL: Category B not filled.")

    # Check if merged cells exist
    if not ws.merged_cells.ranges:
        print("PASS: No merged cells remaining.")
    else:
        print(f"FAIL: Merged cells remain: {ws.merged_cells.ranges}")

    os.remove(filename)

def main():
    source = create_test_file()
    cleaned = excel_cleaner.clean_excel_file(source)

    if cleaned:
        print(f"Cleaner returned: {cleaned}")
        verify_cleaned(cleaned)
        os.remove(source)
    else:
        print("Cleaner returned None (Failed)")

if __name__ == "__main__":
    main()
